import pandas as pd
import sys
import openpyxl
import xlrd
import math
import datetime

# read the file name in the command line
file = sys.argv[1]
path = file[0:-5]

reply = open(path+"反馈.txt",'w+') 

df=pd.read_excel(path+".xlsx", sheet_name=0, converters={'交易卡号':str, '交易账号':str, '客户账号':str,'卡号':str, "账号":str,"帐号":str, "交易日期":str, "交易时间":str, "交易对手账卡号":str, "交易对方账卡号":str, "对方账号":str, "交易柜员号":str, "交易是否成功":str, "柜员号":str, "交易户名":str, "交易流水号":str, "交易网点代码":str, '交易行号':str, '币种号':str})

df.rename(columns=lambda x: x.strip(), inplace=True)

cols = df.columns

list1 = []
list2 = []
list4 = []
boolean = False
row_num = 0
if "交易时间" not in cols and "交易日期" not in cols and "交易卡号" not in cols and "交易账号" not in cols and "交易金额" not in cols and "交易余额" not in cols and "对方户名" not in cols:
    for k in range(len(df)):
        list = []
        for a in range(len(cols)):
            row_value = str(df.iloc[k][a])
            list.append(row_value)
        if "客户编号" in list[0] and "客户账号" in list[0] and "客户名称" in list[0]:
            boolean = True
            str1 = list[0]
            char1 = '客户账号'
            char2 = "，"
            npos1 = int(str1.index(char1))
            npos2 = int(str1.index(char2, npos1))
            start = npos1 + 5
            end = npos2
            acc_num = str1[start:end]
            list1.append(acc_num)
            if row_num != 0:
                list2.append(row_num)
            row_num=0
            list4.append(k)
        elif "根据查询条件" in list[0] or "企业活期明细信息" in list[0]:
            if row_num != 0:
                list2.append(row_num)
            row_num=0
            list4.append(k)
        elif "交易时间" in list or "交易日期" in list or "交易卡号" in list or "交易账号" in list or "交易金额" in list or "交易余额" in list or "对方户名" in list:
            list4.append(k)
            row_num=0
        elif k == len(df)-1:
            if len(list2) != 0:
                list2.append(row_num+1)
        else:
            if boolean == True:
                row_num += 1

if "交易时间" not in cols and "交易日期" not in cols and "交易卡号" not in cols and "交易账号" not in cols and "交易金额" not in cols and "交易余额" not in cols and "对方户名" not in cols:
    drop_num = 0
    for row in range(len(df)):
        list = []
        for col in range(len(cols)):
            row_value = str(df.iloc[row][col])
            list.append(row_value)
        if "交易时间" in list or "交易日期" in list or "交易卡号" in list or "交易账号" in list or "交易金额" in list or "交易余额" in list or "对方户名" in list:
            df.columns = df.loc[row,:]
            cols = df.columns
            df.rename(columns=lambda x: x.strip(), inplace=True)
            df = df.drop(df.index[0:row+1])
            drop_num += row
            if list4 != 0:
                for drop_value in list4:
                    if drop_value > row:
                        df = df.drop(df.index[drop_value-drop_num-1])
                        drop_num += 1
            break

df = df.reset_index(drop=True)

list3 = []
cols = df.columns
if "交易卡号" not in cols:
    df.insert(loc=0, column='交易卡号', value=" ")
if "交易户名" not in cols:
    df.insert(loc=0, column='交易户名', value=" ")
if "交易对方账卡号" not in cols:
    df.insert(loc=0, column='交易对方账卡号', value=" ")
if "交易时间" not in cols:
    if "交易日期" not in cols:
        df.insert(loc=0, column='交易日期', value=" ")
if "交易账号" not in cols:
    if len(list2) != 0:
        df.insert(loc=0, column='交易账号', value=list1[0])
        for list_len in range(len(list2)):
            if list_len == 1:
                start1 = list2[list_len-1]
                end1 = list2[list_len-1]+list2[list_len]
                value = list1[list_len]
                list3.append(end1)
                df.loc[start1:end1 ,["交易账号"]] = value
            if list_len > 1:
                start1 = list3[list_len-2]
                end1 = start1+list2[list_len]
                value = list1[list_len]
                list3.append(end1)
                df.loc[start1:end1 ,["交易账号"]] = value
    else:
        df.insert(loc=0, column='交易账号', value=" ")

cols = df.columns
if "会计日期" in cols:
    df['交易日期'] = pd.to_datetime(df["会计日期"],format='%Y%m%d')
if "交易时间" in cols:
    if "交易日期" in cols:
        char = "000000"
        df['交易日期'] = df['交易日期'].fillna(char)
        df['交易时间'] = df['交易时间'].fillna(char)
        df["交易日期"] = df["交易日期"]+" "+df["交易时间"]
        df["交易日期"].replace('\t','',regex=True, inplace=True)
        df["交易日期"].replace('\.', ':',regex=True, inplace=True)
        df = df.drop(columns=["交易时间"])
if "入帐日期" in cols:
    if "入帐时间" not in cols:
        df["交易日期"] = df["入帐日期"]
        df["交易日期"].replace('\t','',regex=True, inplace=True)
        df["交易日期"].replace('\.', ':',regex=True, inplace=True)
        df = df.drop(columns=["入帐日期"])
if "入帐时间" in cols:
    if "入帐日期" in cols:
        char = "000000"
        df['入帐日期'] = df['入帐日期'].fillna(char)
        df['入帐时间'] = df['入帐时间'].fillna(char)
        df["交易日期"] = df["入帐日期"]+" "+df["入帐时间"]
        df["交易日期"].replace('\t','',regex=True, inplace=True)
        df["交易日期"].replace('\.', ':',regex=True, inplace=True)
        df = df.drop(columns=["入帐日期"])
        df = df.drop(columns=["入帐时间"])
    else:
        df["交易时间"] = df["入帐时间"]
        df["交易时间"].replace('\t','',regex=True, inplace=True)
        df["交易时间"].replace('\.', ':',regex=True, inplace=True)
        df = df.drop(columns=["入帐时间"])
if "交易时间" in cols:
    if "交易日期" not in cols:
        df["交易时间"].replace('\t','',regex=True, inplace=True)
        df["交易时间"].replace('\.', ':',regex=True, inplace=True)
if "交易日期" in cols:
    if "交易时间" not in cols:
        df["交易日期"].replace('\t','',regex=True, inplace=True)
        df["交易日期"].replace('\.', ':',regex=True, inplace=True)
if "交易对方账号" in cols:
    if "交易对方卡号" in cols:
        df["交易对方账卡号"] = df["交易对方账号"]+" "+df["交易对方卡号"]
if "户名" in cols:
    df["交易户名"] = df["户名"]
    df = df.drop(columns=["户名"])
if "卡号" in cols:
    df["交易卡号"] = df["卡号"]
    df = df.drop(columns=["卡号"])
if "账号" in cols:
    df["交易账号"] = df["账号"]
    df = df.drop(columns=["账号"])
if "帐号" in cols:
    df["交易账号"] = df["帐号"]
    df = df.drop(columns=["帐号"])
if "客户账号" in cols:
    df["交易账号"] = df["客户账号"]
    df = df.drop(columns=["客户账号"])

cols = df.columns
for i in range(len(cols)):
    if str(cols[i]) == "交易时间":
        df[["交易时间"]] = df[["交易时间"]].astype(str)
        df['交易时间'] = pd.to_datetime(df['交易时间'], errors='coerce')
        df.sort_values('交易时间', inplace=True, ascending=True)

    elif str(cols[i]) == "交易日期":
        df[["交易日期"]] = df[["交易日期"]].astype(str)
        df['交易日期'] = pd.to_datetime(df['交易日期'], errors='coerce')
        df.sort_values('交易日期', inplace=True, ascending=True)

    elif str(cols[i]) == "交易卡号":
        df[["交易卡号"]] = df[["交易卡号"]].astype(str)

    elif str(cols[i]) == "日志号":
        df[["日志号"]] = df[["日志号"]].astype(str)
        df["日志号"]=[i.replace('.0', '') for i in df["日志号"]]

    elif str(cols[i]) == "交易账号":
        df[["交易账号"]] = df[["交易账号"]].astype(str)

    elif str(cols[i]) == "交易金额":
        df[["交易金额"]] = df[["交易金额"]].astype(str)

    elif str(cols[i]) == "交易余额":
        df[["交易余额"]] = df[["交易余额"]].astype(str)

    elif str(cols[i]) == "发生额":
        df[["发生额"]] = df[["发生额"]].astype(str)

    elif str(cols[i]) == "余额":
        df[["余额"]] = df[["余额"]].astype(str)

    elif str(cols[i]) == "交易后余额":
        df[["交易后余额"]] = df[["交易后余额"]].astype(str)

df.replace('\s+','',regex=True, inplace=True)

cols = df.columns
df = df.fillna('')

cols = df.columns
order = []
for a in range(len(cols)):
    if str(cols[a]) == "交易金额" or str(cols[a]) == "发生额":
        order.append(cols[a])
for b in range(len(cols)):
    if str(cols[b]) == "交易余额" or str(cols[b]) == "余额" or str(cols[b]) == "交易后余额" or str(cols[b]) == "联机余额" or str(cols[b]) == "发生时点余额" or str(cols[b]) == "账户余额":
        order.append(cols[b])
for c in range(len(cols)):
    if str(cols[c]) == "收付标志" or str(cols[c]) == "借贷标志" or str(cols[c]) == "借贷" or "交易方向" in str(cols[c]) or str(cols[c]) == "借贷标识" or str(cols[c]) == "借贷方向":
        order.append(cols[c])
for d in range(len(cols)):
    if str(cols[d]) != "收付标志" and str(cols[d]) != "借贷标志" and str(cols[d]) != "交易余额" and str(cols[d]) != "余额" and str(cols[d]) != "交易金额" and str(cols[d]) != "发生额" and str(cols[d]) != "借贷" and str(cols[d]) != "交易后余额" and str(cols[d]) != "联机余额" and "交易方向" not in str(cols[d]) and str(cols[d]) != "发生时点余额" and str(cols[d]) != "借贷标识" and str(cols[d]) != "账户余额" and str(cols[d]) != "借贷方向":        order.append(cols[d])
df = df[order]

jg_list = df["交易卡号"].unique()
jg_list1 = df["交易账号"].unique()
jg_list3 = df["交易户名"].unique()

num = 0
non_list = ["nan"," ",""]
new_wb = pd.ExcelWriter(path+'拆分表.xlsx')

for jg in jg_list1:
    warning = 0
    sig_bool = False
    child_wb = df[df['交易账号'] == jg]
    jg_list2 = child_wb["交易卡号"].unique()
    warning_list = []
    for jg1 in jg_list:
        if jg1 in jg_list2:
            warning_list.append(jg1)
            warning += 1
    if warning >= 2:
        for sig in warning_list:
            if sig in non_list:
                sig_bool = True
        if sig_bool == True:
            if str(jg) == " " or str(jg) == "":
                print("---------------------"+"交易账号：nan拆分警告"+"---------------------"+"\n"+"包含卡号：",file=reply)
            else:
                print("---------------------"+"交易账号："+jg+"拆分警告"+"---------------------"+"\n"+"包含卡号：",file=reply)
            for sig_value in warning_list:
                if str(sig_value) == " " or str(sig_value) == "":
                    print("nan",file=reply)
                else:
                    print(sig_value,file=reply)

for jg in jg_list:
    warning = 0
    sig_bool1 = False
for jg in jg_list:
    warning1 = 0
    sig_bool1 = False
    child_wb = df[df['交易卡号'] == jg]
    jg_list2 = child_wb["交易账号"].unique()
    warning_list1 =[]
    for jg1 in jg_list1:
        if jg1 in jg_list2:
            warning_list1.append(jg1)
            warning1 += 1
    if warning1 >= 2:
        for sig in warning_list1:
            if sig in non_list:
                sig_bool1 = True
        if sig_bool1 == True:
            if str(jg) == " " or str(jg) == "":
                print("---------------------"+"交易账号：nan拆分警告"+"---------------------"+"\n"+"包含卡号：",file=reply)
            else:
                print("---------------------"+"交易卡号："+jg+"拆分警告"+"---------------------"+"\n"+"包含账号：",file=reply)
            for sig_value in warning_list1:
                if str(sig_value) == " " or str(sig_value) == "":
                    print("nan",file=reply)
                else:
                    print(sig_value,file=reply)

for jg in jg_list:
    child_wb = df[df['交易卡号'] == jg]
    jg_list2 = child_wb["交易账号"].unique()
    for jg1 in jg_list1:
        if jg1 in jg_list2:
            child_wb1 = child_wb[child_wb['交易账号'] == jg1]
            jg_list4 = child_wb1["交易户名"].unique()
            for jg2 in jg_list3:
                if jg2 in jg_list4:
                    num += 1
                    child_wb2 = child_wb1[child_wb1['交易户名'] == jg2]
                    if jg not in non_list or jg1 not in non_list or jg2 not in non_list:
                        child_wb2.to_excel(new_wb, index=False, sheet_name=str(jg2)+' '+str(jg)[-4:]+' '+str(jg1)[-4:])
                        print("拆分出第"+str(num)+"个工作表："+str(jg2)+' '+str(jg)[-4:]+' '+str(jg1)[-4:],file=reply)
                    else:
                        child_wb2.to_excel(new_wb, index=False, sheet_name='nan')
                        print("拆分出第"+str(num)+"个工作表："+'nan',file=reply)
new_wb.save()   
    
print("共有"+ str(num)+"个工作表",file=reply)

def xldate_to_datetime(xldatetime): #something like 43705.6158241088

    tempDate = datetime.datetime(1899, 12, 30)
    (days, portion) = math.modf(xldatetime)

    deltaDays = datetime.timedelta(days=days)
    secs = int(24 * 60 * 60 * portion)
    detlaSeconds = datetime.timedelta(seconds=secs)
    TheTime = (tempDate + deltaDays + detlaSeconds )
    return TheTime.strftime("%Y-%m-%d %H:%M:%S")

wb = xlrd.open_workbook(path+"拆分表.xlsx")

oriwb=openpyxl.load_workbook(path+".xlsx",read_only=True)
oriws = oriwb[oriwb.sheetnames[0]]

sheet_num = wb.nsheets
print("工作表数量：" + str(sheet_num))
sheet_name = wb.sheet_names()


for j in range(sheet_num):
    book = openpyxl.Workbook()
    sh = book.active
    sh.title = sheet_name[j]+"整理"
    orish = book.create_sheet(title='原始数据',index=1)
    original_data = wb.sheet_by_index(j)

    print("工作表：" + str(sheet_name[j]))
    nrows = original_data.nrows # calc the number of rows
    print("行数：" + str(nrows))
    ncols = original_data.ncols # calc the number of columns
    print("列数：" + str(ncols))

    for m,row in enumerate(oriws.iter_rows()):
        for n,cell in enumerate(row):
            orish.cell(row=m+1, column=n+1, value=cell.value)

    other = 0
    currency_num = 0
    for i in range(ncols):
        original_name = str(original_data.cell(0, i))
        new_name = original_name[6:-1]
        if new_name == "交易卡号":
            card_num = original_data.col(i)
            for curr_row in range(nrows):
                original_card_num = str(card_num[curr_row])
                if original_card_num[0:5] == "empty":
                    sh.cell(curr_row+1, 1, " ")
                elif original_card_num[0:4] == "text":
                    new_card_num = original_card_num[6:-1]
                    if str(new_card_num) == "nan":
                        sh.cell(curr_row+1, 1, " ")
                    else:
                        sh.cell(curr_row+1, 1, str(new_card_num))

        elif new_name == "交易账号":
            acc_num = original_data.col(i)
            for curr_row in range(nrows):
                original_acc_num = str(acc_num[curr_row])
                if original_acc_num[0:5] == "empty":
                    sh.cell(curr_row, 1, " ")
                elif original_acc_num[0:4] == "text":
                    new_acc_num = original_acc_num[6:-1]
                    if new_acc_num == "nan":
                        sh.cell(curr_row+1, 2, " ")
                    else:
                        sh.cell(curr_row+1, 2, str(new_acc_num))

        elif new_name == "交易户名" or new_name == "户名":
            acc_name = original_data.col(i)
            for curr_row in range(nrows):
                original_acc_name = str(acc_name[curr_row])
                if original_acc_name[0:5] == "empty":
                    sh.cell(curr_row+1, 4, " ")
                elif original_acc_name[0:4] == "text":
                    new_acc_name = str(original_acc_name[6:-1])
                    sh.cell(curr_row+1, 4, str(new_acc_name))

        elif new_name == "交易金额" or new_name == "发生额":
            amount = original_data.col(i)
            for curr_row in range(nrows):
                original_amount = str(amount[curr_row])
                if original_amount[0:4] == "text":
                    new_amount = original_amount[6:-1].replace(",","")
                    if new_amount == "'":
                        sh.cell(curr_row+1, 7, 0)
                    elif new_amount != "'" and curr_row != 0:
                        float_amount = float(new_amount)
                        sh.cell(curr_row+1, 7, float_amount)
                        cell_x = sh.cell(curr_row+1, 7).value
                        sh.cell(curr_row+1, 7, float(cell_x) + int(0)).number_format = '#,##0.00'
                    elif new_amount == "nan":
                        sh.cell(curr_row+1, 7, int(0))
                elif original_amount[0:6] == "number":
                    new_amount = original_amount[7:]
                    if str(new_amount) == "'":
                        sh.cell(curr_row+1, 7, 0)
                    if str(new_amount) != "'" and curr_row != 0:
                        float_amount = float(new_amount)
                        sh.cell(curr_row+1, 7, format(float_amount, '.2f'))
                        cell_x = sh.cell(curr_row+1, 7).value
                        sh.cell(curr_row+1, 7).value = float(cell_x) + int(0)
            
        elif new_name == "收付标志" or new_name == "借贷标志" or new_name == "借贷" or new_name == "借贷标识" or new_name == "借贷方向" or "交易方向" in new_name:
            currency_num += 1
            expence_num = 0
            nega_num = 0
            trans_value = original_data.col(i-2)
            for curr_row in range(nrows):
                label = str(original_data.cell(curr_row, i))[6:-1]
                if label == "出" or label == "支出" or label == "借" or "借" in label or "付" in label or "D" in label:
                    original_expence = str(trans_value[curr_row])
                    if original_expence[0:4] == "text":
                        new_expence = original_expence[6:-1]
                    if original_expence[0:6] == "number":
                        new_expence = original_expence[7:]
                    expence_num += 1
                    if str(new_expence[0]) == "-":
                        nega_num += 1
            balance =  original_data.col(i-1)
            recapture = original_data.col(i)
            for curr_row in range(nrows):
                original_balance = str(balance[curr_row])
                if original_balance[0:4] == "text":
                    if curr_row != 0:
                        new_balance = original_balance[6:-1].replace(",","")
                        float_balance = float(new_balance)
                        sh.cell(curr_row+1, 11, float_balance)
                        cell_x = sh.cell(curr_row+1, 11).value
                        sh.cell(curr_row+1, 11, float(cell_x) + int(0)).number_format = '#,##0.00'
                elif original_balance[0:6] == "number":
                    if curr_row != 0:
                        new_balance = original_balance[7:]
                        sh.cell(curr_row+1, 11, new_balance)
                        cell_x = sh.cell(curr_row+1, 11).value
                        sh.cell(curr_row+1, 11).value = float(cell_x) + int(0)
                
                original_recapture = str(recapture[curr_row])
                if original_recapture[0:5] == "empty":
                    sh.cell(curr_row+1, 14, " ")
                elif original_recapture[0:4] == "text":
                    new_recapture = original_recapture[6:-1]
                    sh.cell(curr_row+1, 14, new_recapture)
                
                label = str(original_data.cell(curr_row, i))[6:-1]
                
                if label == "进" or label == "收入" or label == "贷" or "贷" in label or "收" in label or "C" in label:
                    original_income = str(trans_value[curr_row])
                    if original_income[0:4] == "text":
                        new_income = original_income[6:-1].replace(",","")
                        if curr_row != 0:
                            float_income = float(new_income)
                            sh.cell(curr_row+1, 8,float_income)
                            cell_x = sh.cell(curr_row+1, 8).value
                            sh.cell(curr_row+1, 8, float(cell_x) + int(0)).number_format = '#,##0.00'
                            sh.cell(curr_row+1, 9, int(0))
                            sh.cell(curr_row+1, 10, float_income)
                            cell_y = sh.cell(curr_row+1, 10).value
                            sh.cell(curr_row+1, 10, float(cell_y) + int(0)).number_format = '#,##0.00'
                    elif original_income[0:6] == "number":
                        new_income = original_income[7:]
                        if curr_row != 0:
                            float_income = float(new_income)
                            sh.cell(curr_row+1, 8, float_income)
                            cell_x = sh.cell(curr_row+1, 8).value
                            sh.cell(curr_row+1, 8, float(cell_x) + int(0)).number_format = '#,##0.00'
                            sh.cell(curr_row+1, 9, int(0))
                            sh.cell(curr_row+1, 10, float_income)
                            cell_y = sh.cell(curr_row+1, 10).value
                            sh.cell(curr_row+1, 10, float(cell_y) + int(0)).number_format = '#,##0.00'


                if label == "出" or label == "支出" or label == "借" or "借" in label or "付" in label or "D" in label:
                    original_expence = str(trans_value[curr_row])
                    if original_expence[0:4] == "text":
                        new_expence = original_expence[6:-1].replace(",","")
                        if (expence_num-expence_num*0.1) < nega_num < (expence_num+expence_num*0.1):
                            pos_expence = new_expence[1:]
                            if curr_row != 0:
                                float_expence = float(pos_expence)
                                sh.cell(curr_row+1, 9, float_expence)
                                cell_x = sh.cell(curr_row+1, 9).value
                                sh.cell(curr_row+1, 9, float(cell_x) + int(0)).number_format = '#,##0.00'
                                sh.cell(curr_row+1, 8, int(0))
                                sh.cell(curr_row+1, 10, 0-float(format(float_expence, '.2f')))
                                cell_y = sh.cell(curr_row+1, 10).value
                                sh.cell(curr_row+1, 10, float(cell_y) + int(0)).number_format = '#,##0.00'
                        else:
                            if curr_row != 0:
                                float_expence = float(new_expence)
                                sh.cell(curr_row+1, 9, float_expence)
                                cell_x = sh.cell(curr_row+1, 9).value
                                sh.cell(curr_row+1, 9, float(cell_x) + int(0)).number_format = '#,##0.00'
                                sh.cell(curr_row+1, 8, int(0))
                                sh.cell(curr_row+1, 10, 0-float(format(float_expence, '.2f')))
                                cell_y = sh.cell(curr_row+1, 10).value
                                sh.cell(curr_row+1, 10, float(cell_y) + int(0)).number_format = '#,##0.00'
                    if original_expence[0:6] == "number":
                        new_expence = original_expence[7:]
                        if (expence_num-expence_num*0.1) < nega_num < (expence_num+expence_num*0.1):
                            pos_expence = new_expence[1:]
                            if curr_row != 0:
                                float_expence = float(pos_expence)
                                sh.cell(curr_row+1, 9, float_expence)
                                cell_x = sh.cell(curr_row+1, 9).value
                                sh.cell(curr_row+1, 9, float(cell_x) + int(0)).number_format = '#,##0.00'
                                sh.cell(curr_row+1, 8, int(0))
                                sh.cell(curr_row+1, 10, 0-float(format(float_expence, '.2f')))
                                cell_y = sh.cell(curr_row+1, 10).value
                                sh.cell(curr_row+1, 10, float(cell_y) + int(0)).number_format = '#,##0.00'
                        else:
                            if curr_row != 0:
                                float_expence = float(new_expence)
                                sh.cell(curr_row+1, 9, float_expence)
                                cell_x = sh.cell(curr_row+1, 9).value
                                sh.cell(curr_row+1, 9, float(cell_x) + int(0)).number_format = '#,##0.00'
                                sh.cell(curr_row+1, 8, int(0))
                                sh.cell(curr_row+1, 10, 0-float(format(float_expence, '.2f')))
                                cell_y = sh.cell(curr_row+1, 10).value
                                sh.cell(curr_row+1, 10, float(cell_y) + int(0)).number_format = '#,##0.00'

                if label == "'":
                    sh.cell(curr_row+1, 8, int(0))
                    sh.cell(curr_row+1, 9, int(0))
                    sh.cell(curr_row+1, 10, int(0))

                if curr_row == 1:
                    sh.cell(2, 12, new_balance).number_format = '#,##0.00'
                    cell_x = sh.cell(2,12).value
                    sh.cell(2, 12).value = float(cell_x) + int(0)
                    cell_a = sh.cell(curr_row+1, 11).value
                    cell_b = sh.cell(curr_row+1, 12).value
                    if str(type(cell_a))[8:-2] == "NoneType":
                        cell_a = "nan"
                    if str(type(cell_a))[8:-2] == "NoneType":
                        cell_b = "nan"
                    sh.cell(2, 13).value = float(cell_a) - float(cell_b)
                    sh.cell(2, 13).number_format = '#,##0.00'
                if curr_row > 1:
                    cell_c = sh.cell(curr_row, 12).value
                    cell_d = sh.cell(curr_row+1, 8).value
                    cell_e = sh.cell(curr_row+1, 9).value
                    if str(type(cell_c))[8:-2] == "NoneType":
                        cell_c = "nan"
                    if str(type(cell_d))[8:-2] == "NoneType":
                        cell_d = "nan"
                    if str(type(cell_e))[8:-2] == "NoneType":
                        cell_e = "nan"
                    sh.cell(curr_row+1, 12).value = float(cell_c) + float(cell_d) - float(cell_e)
                    sh.cell(curr_row+1, 12).number_format = '#,##0.00'
                    cell_a = sh.cell(curr_row+1, 11).value
                    cell_b = sh.cell(curr_row+1, 12).value
                    if str(type(cell_a))[8:-2] == "NoneType":
                        cell_a = "nan"
                    if str(type(cell_a))[8:-2] == "NoneType":
                        cell_b = "nan"
                    sh.cell(curr_row+1, 13).value = float(cell_a) - float(cell_b)
                    sh.cell(curr_row+1, 13).number_format = '#,##0.00'

        elif new_name == "贷方发生额":
            currency_num += 1
            income = original_data.col(i)
            for curr_row in range(nrows):
                original_income = str(income[curr_row])
                if original_income[0:4] == "text":
                    new_income = original_income[6:-1].replace(",","")
                    if curr_row != 0:
                        float_income = float(new_income)
                        sh.cell(curr_row+1, 8, float_income)
                        cell_x = sh.cell(curr_row+1, 8).value
                        sh.cell(curr_row+1, 8, float(cell_x) + int(0)).number_format = '#,##0.00'
                        if float_income != 0:
                            sh.cell(curr_row+1, 10, float_income)
                            cell_y = sh.cell(curr_row+1, 10).value
                            sh.cell(curr_row+1, 10, float(cell_y) + int(0)).number_format = '#,##0.00'
                            sh.cell(curr_row+1, 7, float_income)
                            cell_z = sh.cell(curr_row+1, 7).value
                            sh.cell(curr_row+1, 7, float(cell_z) + int(0)).number_format = '#,##0.00'
                if original_income[0:6] == "number":
                    new_income = original_income[7:]
                    if curr_row != 0:
                        float_income = float(new_income)
                        sh.cell(curr_row+1, 8, float_income)
                        cell_x = sh.cell(curr_row+1, 8).value
                        sh.cell(curr_row+1, 8, float(cell_x) + int(0)).number_format = '#,##0.00'
                        if float_income != 0:
                            sh.cell(curr_row+1, 10, float_income)
                            cell_y = sh.cell(curr_row+1, 10).value
                            sh.cell(curr_row+1, 10, float(cell_y) + int(0)).number_format = '#,##0.00'
                            sh.cell(curr_row+1, 7, float_income)
                            cell_z = sh.cell(curr_row+1, 7).value
                            sh.cell(curr_row+1, 7, float(cell_z) + int(0)).number_format = '#,##0.00'

        elif new_name == "借方发生额":
            currency_num += 1
            expence = original_data.col(i)
            num = 0
            nega = 0
            for curr_row in range(nrows):
                original_expence = str(expence[curr_row])
                if original_expence[0:4] == "text":
                     new_expence = original_expence[6:-1].replace(",","")
                     if curr_row != 0:
                        float_expence = float(new_expence)
                        if str(new_expence[0]) == "-":
                            nega += 1
                        if float_expence != 0:
                            num += 1
                elif original_expence[0:6] == "number":
                    new_expence = original_expence[7:]
                    if curr_row != 0:
                        float_expence = float(new_expence)
                        if str(new_expence[0]) == "-":
                            nega += 1
                        if float_expence != 0:
                            num += 1
            for curr_row in range(nrows):
                original_expence = str(expence[curr_row])
                if original_expence[0:4] == "text":
                    new_expence = original_expence[6:-1].replace(",","")
                    if (num-num*0.1) < nega < (num+num*0.1):
                        if str(new_expence[0]) == "-":
                            pos_expence = new_expence[1:]
                            if curr_row != 0:
                                float_expence = float(pos_expence)
                                sh.cell(curr_row+1, 9, float_expence)
                                cell_x = sh.cell(curr_row+1, 9).value
                                sh.cell(curr_row+1, 9, float(cell_x) + int(0)).number_format = '#,##0.00'
                                if float_expence != 0:
                                    sh.cell(curr_row+1, 10, 0-float(format(float_expence, '.2f')))
                                    cell_y = sh.cell(curr_row+1, 10).value
                                    sh.cell(curr_row+1, 10, float(cell_y) + int(0)).number_format = '#,##0.00'
                                    sh.cell(curr_row+1, 7, float(format(float_expence, '.2f')))
                                    cell_z = sh.cell(curr_row+1, 7).value
                                    sh.cell(curr_row+1, 7, float(cell_z) + int(0)).number_format = '#,##0.00'
                    else:
                        if curr_row != 0:
                            float_expence = float(new_expence)
                            sh.cell(curr_row+1, 9, float_expence)
                            cell_x = sh.cell(curr_row+1, 9).value
                            sh.cell(curr_row+1, 9, float(cell_x) + int(0)).number_format = '#,##0.00'
                            if float_expence != 0:
                                sh.cell(curr_row+1, 10, 0-float(format(float_expence, '.2f')))
                                cell_y = sh.cell(curr_row+1, 10).value
                                sh.cell(curr_row+1, 10, float(cell_y) + int(0)).number_format = '#,##0.00'
                                sh.cell(curr_row+1, 7, float_expence)
                                cell_z = sh.cell(curr_row+1, 7).value
                                sh.cell(curr_row+1, 7, float(cell_z) + int(0)).number_format = '#,##0.00'
                elif original_expence[0:6] == "number":
                    new_expence = original_expence[7:]
                    if (num-num*0.1) < nega < (num+num*0.1):
                        if str(new_expence[0]) == "-":
                            pos_expence = new_expence[1:]
                            if curr_row != 0:
                                float_expence = float(pos_expence)
                                sh.cell(curr_row+1, 9, float_expence)
                                cell_x = sh.cell(curr_row+1, 9).value
                                sh.cell(curr_row+1, 9, float(cell_x) + int(0)).number_format = '#,##0.00'
                                if float_expence != 0:
                                    sh.cell(curr_row+1, 10, 0-float(format(float_expence, '.2f')))
                                    cell_y = sh.cell(curr_row+1, 10).value
                                    sh.cell(curr_row+1, 10, float(cell_y) + int(0)).number_format = '#,##0.00'
                                    sh.cell(curr_row+1, 7, float(format(float_expence, '.2f')))
                                    cell_z = sh.cell(curr_row+1, 7).value
                                    sh.cell(curr_row+1, 7, float(cell_z) + int(0)).number_format = '#,##0.00'
                    else:
                        if curr_row != 0:
                            float_expence = float(new_expence)
                            sh.cell(curr_row+1, 9, float_expence)
                            cell_x = sh.cell(curr_row+1, 9).value
                            sh.cell(curr_row+1, 9, float(cell_x) + int(0)).number_format = '#,##0.00'
                            if float_expence != 0:
                                sh.cell(curr_row+1, 10, 0-float(format(float_expence, '.2f')))
                                cell_y = sh.cell(curr_row+1, 10).value
                                sh.cell(curr_row+1, 10, float(cell_y) + int(0)).number_format = '#,##0.00'
                                sh.cell(curr_row+1, 7, float_expence)
                                cell_z = sh.cell(curr_row+1, 7).value
                                sh.cell(curr_row+1, 7, float(cell_z) + int(0)).number_format = '#,##0.00'

        elif new_name == "账户余额" or new_name == "本笔余额" or new_name == "交易后余额" or new_name == "联机余额" or new_name == "发生时点余额":
            balance =  original_data.col(i)
            for curr_row in range(nrows):
                original_balance = str(balance[curr_row])
                if original_balance[0:4] == "text":
                    new_balance = original_balance[6:-1].replace(",","")
                    if curr_row != 0:
                        float_balance = float(new_balance)
                        sh.cell(curr_row+1, 11, float_balance)
                        cell_x = sh.cell(curr_row+1, 11).value
                        sh.cell(curr_row+1, 11, float(cell_x) + int(0)).number_format = '#,##0.00'
                elif original_balance[0:6] == "number":
                    new_balance = original_balance[7:]
                    if curr_row != 0:
                        float_balance = float(new_balance)
                        sh.cell(curr_row+1, 11, float_balance)
                        cell_x = sh.cell(curr_row+1, 11).value
                        sh.cell(curr_row+1, 11, float(cell_x) + int(0)).number_format = '#,##0.00'
                if curr_row == 1:
                    sh.cell(2, 12, new_balance)
                    cell_x = sh.cell(2, 12).value
                    sh.cell(2, 12, float(cell_x) + int(0)).number_format = '#,##0.00'
                    cell_a = sh.cell(2, 11).value
                    cell_b = sh.cell(2, 12).value
                    if str(type(cell_a))[8:-2] == "NoneType":
                        cell_a = "nan"
                    if str(type(cell_a))[8:-2] == "NoneType":
                        cell_b = "nan"
                    sh.cell(2, 13).value = float(cell_a) - float(cell_b)
                    sh.cell(2, 13).number_format = '#,##0.00'
                elif curr_row > 1:
                    cell_c = sh.cell(curr_row, 12).value
                    cell_d = sh.cell(curr_row+1, 8).value
                    cell_e = sh.cell(curr_row+1, 9).value
                    if str(type(cell_c))[8:-2] == "NoneType":
                        cell_c = "nan"
                    if str(type(cell_d))[8:-2] == "NoneType":
                        cell_d = "nan"
                    if str(type(cell_e))[8:-2] == "NoneType":
                        cell_e = "nan"
                    sh.cell(curr_row+1, 12).value = float(cell_c) + float(cell_d) - float(cell_e)
                    sh.cell(curr_row+1, 12).number_format = '#,##0.00'
                    cell_a = sh.cell(curr_row+1, 11).value
                    cell_b = sh.cell(curr_row+1, 12).value
                    if str(type(cell_a))[8:-2] == "NoneType":
                        cell_a = "nan"
                    if str(type(cell_a))[8:-2] == "NoneType":
                        cell_b = "nan"
                    sh.cell(curr_row+1, 13).value = float(cell_a) - float(cell_b)
                    sh.cell(curr_row+1, 13).number_format = '#,##0.00'

        elif new_name == "日志号":
            journum = original_data.col(i)
            for curr_row in range(nrows):
                original_journum = str(journum[curr_row])
                if original_journum[0:5] == "empty":
                    sh.cell(curr_row+1, 15, " ")
                elif original_journum[0:4] == "text":
                    new_journum = original_journum[6:-1]
                    if str(new_journum) == 'nan':
                        sh.cell(curr_row+1, 15, '')
                    else:
                        sh.cell(curr_row+1, 15, new_journum)
                elif original_journum[0:6] == "number":
                    new_journum = original_journum[7:]
                    sh.cell(curr_row+1, 15, new_journum)
            
        elif new_name == "凭证种类":
            voucher_cate = original_data.col(i)
            for curr_row in range(nrows):
                original_voucher_cate = str(voucher_cate[curr_row])
                if original_voucher_cate[0:5] == "empty":
                    sh.cell(curr_row+1, 16, " ")
                elif original_voucher_cate[0:4] == "text":
                    new_voucher_cate = original_voucher_cate[6:-1]
                    sh.cell(curr_row+1, 16, new_voucher_cate)
                elif original_voucher_cate[0:6] == "number":
                    new_voucher_cate = original_voucher_cate[7:]
                    if str(new_voucher_cate[-2:]) == ".0":
                        int_voucher_cate = new_voucher_cate[:-2]
                        four_num = str((4-len(int_voucher_cate))*"0"+int_voucher_cate)
                        sh.cell(curr_row+1, 16, four_num)
                    else:
                        sh.cell(curr_row+1, 16, new_voucher_cate)

        elif new_name == "交易时间" or new_name == "交易日期":
            date_list = original_data.col(i)
            for curr_row in range(nrows):
                original_date = str(date_list[curr_row])
                if curr_row != 0:
                    if original_date[0:6] == "xldate":
                        xldate = original_date[7:]
                        new_datetime = xldate_to_datetime(float(xldate))
                        sh.cell(curr_row+1, 6, new_datetime).number_format = 'mm/dd/yyyy;@'
                    else:
                        new_date = original_date[6:-1]
                        sh.cell(curr_row+1, 6, new_date).number_format = 'mm/dd/yyyy;@'

        elif new_name == "查询反馈结果原因":
            feedback = original_data.col(i)
            for curr_row in range(nrows):
                original_feedback = str(feedback[curr_row])
                new_feedback = original_feedback[6:-1]
                if original_feedback[0:5] == "empty":
                    sh.cell(curr_row+1, 3, " ")
                else:
                    if str(new_feedback) == "该账号无交易记录" or str(new_feedback) == "该指令入库时间超时，未能处理，请重新发起" or "成功,查询超限，只查询" in str(new_feedback):   
                        sh.cell(curr_row+1, 3, str(new_feedback))
                        sh.cell(curr_row+1, 7, value="")
                        sh.cell(curr_row+1, 8, value="")
                        sh.cell(curr_row+1, 9, value="")
                        sh.cell(curr_row+1, 10, value="")
                        sh.cell(curr_row+1, 11, value="")
                        sh.cell(curr_row+1, 12, value="")
                        sh.cell(curr_row+1, 13, value="")
                    else:
                        sh.cell(curr_row+1, 3, str(new_feedback))

        elif new_name == "交易是否成功":
            feedback1 = original_data.col(i)
            for curr_row in range(nrows):
                original_feedback1 = str(feedback1[curr_row])
                new_feedback1 = original_feedback1[6:-1]
                if original_feedback1[0:5] == "empty":
                    sh.cell(curr_row+1, 5, " ")
                else:
                    if str(new_feedback1) == "该账号无交易记录" or str(new_feedback1) == "该指令入库时间超时，未能处理，请重新发起" or "成功,查询超限，只查询" in str(new_feedback1):   
                        sh.cell(curr_row+1, 5, str(new_feedback))
                        sh.cell(curr_row+1, 7, value="")
                        sh.cell(curr_row+1, 8, value="")
                        sh.cell(curr_row+1, 9, value="")
                        sh.cell(curr_row+1, 10, value="")
                        sh.cell(curr_row+1, 11, value="")
                        sh.cell(curr_row+1, 12, value="")
                        sh.cell(curr_row+1, 13, value="")
                    else:
                        sh.cell(curr_row+1, 5, str(new_feedback1))

        else:
            if new_name != "余额" and new_name != "交易余额" and new_name != "交易后余额":
                other += 1
                other_col = 16 + other
                other_value = original_data.col(i)
                for curr_row in range(nrows):
                    original_value = str(other_value[curr_row])
                    if original_value[0:5] == "empty":
                        sh.cell(curr_row+1, other_col, " ")
                    elif original_value[0:4] == "text":
                        new_value = original_value[6:-1]
                        sh.cell(curr_row+1, other_col, new_value)
                    elif original_value[0:6] == "number":
                        new_value = original_value[7:]
                        if str(new_value[-2:]) == ".0":
                            int_value = new_value[:-2]
                            sh.cell(curr_row+1, other_col, int_value)
                        else:
                            sh.cell(curr_row+1, other_col, new_value)

    if currency_num == 0:
        for curr_row in range(nrows):
            if curr_row != 0:
                cell_x = sh.cell(curr_row+1, 7).value
                sh.cell(curr_row+1, 10, cell_x)
                cell_z = sh.cell(curr_row+1, 10).value
                sh.cell(curr_row+1, 10, float(cell_z) + int(0)).number_format = '#,##0.00'
                if str(cell_x)[0] == "-":
                    new_value = float(str(cell_x)[1:])
                    sh.cell(curr_row+1, 9, new_value)
                    cell_y = sh.cell(curr_row+1, 9).value
                    sh.cell(curr_row+1, 9, float(cell_y) + int(0)).number_format = '#,##0.00'
                    sh.cell(curr_row+1, 8, int(0))
                else:
                    sh.cell(curr_row+1, 8, cell_x)
                    cell_y = sh.cell(curr_row+1, 8).value
                    sh.cell(curr_row+1, 8, float(cell_y) + int(0)).number_format = '#,##0.00'
                    sh.cell(curr_row+1, 9, int(0))
                if curr_row > 1:
                    cell_c = sh.cell(curr_row, 12).value
                    cell_d = sh.cell(curr_row+1, 8).value
                    cell_e = sh.cell(curr_row+1, 9).value
                    if str(type(cell_c))[8:-2] == "NoneType":
                        cell_c = "nan"
                    if str(type(cell_d))[8:-2] == "NoneType":
                        cell_d = "nan"
                    if str(type(cell_e))[8:-2] == "NoneType":
                        cell_e = "nan"
                    sh.cell(curr_row+1, 12).value = float(cell_c) + float(cell_d) - float(cell_e)
                    sh.cell(curr_row+1, 12).number_format = '#,##0.00'
                    cell_a = sh.cell(curr_row+1, 11).value
                    cell_b = sh.cell(curr_row+1, 12).value
                    if str(type(cell_a))[8:-2] == "NoneType":
                        cell_a = "nan"
                    if str(type(cell_a))[8:-2] == "NoneType":
                        cell_b = "nan"
                    sh.cell(curr_row+1, 13).value = float(cell_a) - float(cell_b)
                    sh.cell(curr_row+1, 13).number_format = '#,##0.00'


    sh.cell(1, 1, "交易卡号")
    sh.cell(1, 2, "交易账号")
    sh.cell(1, 3, "查询反馈结果原因")
    sh.cell(1, 4, "交易户名")
    sh.cell(1, 5, "交易是否成功")
    sh.cell(1, 6, "交易日期")
    sh.cell(1, 7, "交易金额")
    sh.cell(1, 8, "收入")
    sh.cell(1, 9, "支出")
    sh.cell(1, 10, "净流")
    sh.cell(1, 11, "余额")
    sh.cell(1, 12, "公式余额")
    sh.cell(1, 13, "公式校验")
    sh.cell(1, 14, "收付标志")
    sh.cell(1, 15, "日志号")
    sh.cell(1, 16, "凭证种类")

    title = path[:10]
    saveExcel =str(path+sheet_name[j])+".xlsx"
    book.save(saveExcel)




